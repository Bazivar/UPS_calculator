import openpyxl
from openpyxl.chart import LineChart, Reference
from calc import UPS_calc, UPS_unrounded_calc
from openpyxl.writer.excel import save_workbook
import webbrowser

#initial parameters


def excel_create(model, power, battery_cap, battery_num, ext_model, ext_bat_num, ext_bat_cap, ext_num, efficiency, max_power):

    FILE_NAME = model.replace(' ', '_') + '.xlsx'

    #creating an Excel file

    wb = openpyxl.load_workbook(filename='template.xlsx')

    # #deleting the default sheet
    # sheet_name = wb.sheetnames[0]
    # sheet = wb.get_sheet_by_name(sheet_name)
    # wb.remove_sheet(sheet)

    # #creating new sheets
    # wb.create_sheet('Расчёт')
    # wb.create_sheet('Данные')

    sheet = wb['Расчёт']
    data = wb['Данные']

    #writing the parameters
    # sheet['A1'] = 'Модель: '
    sheet['A2'] = model
    # sheet['A3'] = 'Потребляемая мощность: '
    sheet['A4'] = str(power) + ' Вт'
    # sheet['A5'] = 'Ёмкость одной АКБ в ИБП: '
    sheet['A6'] = str(battery_cap) + ' Ач'
    # sheet['A7'] = 'Число АКБ в ИБП: '
    sheet['A8'] = str(battery_num) + ' шт'
    # sheet['A9'] = 'Аккумуляторное расширение: '
    sheet['A10'] = ext_model
    # sheet['A11'] = 'Число АКБ в аккумуляторном расширении: '
    sheet['A12'] = str(ext_bat_num) + ' шт'
    # sheet['A13'] = 'Ёмкость одной АКБ в аккумуляторном расширении: '
    sheet['A14'] = str(ext_bat_cap) + ' Ач'
    # sheet['A15'] = 'Число аккумуляторных расширений: '
    sheet['A16'] = str(ext_num) + ' шт'
    # sheet['A17'] = 'Максимальная мощность ИБП: '
    sheet['A18'] = str(max_power) + ' Вт'

    #calculating UPS back-up time
    result = UPS_calc(power, battery_cap, battery_num, ext_bat_num, ext_bat_cap, ext_num, efficiency)
    sheet['A21'] = str(result) + ' мин'


    #making the data table
    data['B1'] = 'Время автономной работы, мин'
    for i in range (1,max_power+1):
        data[f'A{i+1}'] = i
        data[f'B{i+1}'] = UPS_unrounded_calc(i, battery_cap, battery_num, ext_bat_num, ext_bat_cap, ext_num, efficiency)

    #creating the chart
    chart = LineChart()
    chart.title = "График времени автономной работы " + model
    chart.style = 8
    chart.width = 30
    chart.height = 10
    chart.x_axis.title = "Нагрузка, Вт"
    chart.y_axis.title = "Время, мин"
    chart.y_axis.scaling.logBase = 2
    chart_data = Reference(data, min_col=2, min_row=50, max_row=max_power+1)
    chart.add_data(chart_data, titles_from_data=True)
    sheet.add_chart(chart, 'C2')

    #saving the Excel file
    save_workbook(wb, FILE_NAME)

    #opening the Excel file on the PC
    webbrowser.open(FILE_NAME)

